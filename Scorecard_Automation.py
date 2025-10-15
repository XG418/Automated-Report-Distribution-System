import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string

SCORECARD_PATH = "sample_data/Scorecard_template.xlsx"
SOURCE_PATH = "sample_data/Source_demo.xlsx"
OUTPUT_DIR = "output"
SCORECARD_SHEET_NAME = "Scorecard"
FACILITY_COL_LETTER = "J"
REGION_COL_LETTER = "K"


METRICS = [
    {"row": 5,  "facility": "B10", "region": "C10", "sheet": "Sheet1"},
    {"row": 7,  "facility": "B16", "region": "C16", "sheet": "Sheet1"},
    {"row": 8,  "facility": "B30", "region": "C30", "sheet": "Sheet1", "multiply": 100},
    {"row": 9,  "facility": "B23", "region": "C23", "sheet": "Sheet2"},
    {"row": 10, "facility": "B36", "region": "C36", "sheet": "Sheet2", "multiply": 100},
    {"row": 11, "facility": "B25", "region": "C25", "sheet": "Sheet3", "multiply": 100},
    {"row": 13, "facility": "B22", "region": "C22", "sheet": "Sheet1"},
    {"row": 14, "facility": "B27", "region": "C27", "sheet": "Sheet3"},
    {"row": 20, "facility": "B5",  "region": "C5",  "sheet": "Sheet1"},
    {"row": 21, "facility": "B19", "region": "C19", "sheet": "Sheet2"},
]

def _ensure_demo_files():
    """Create sample Scorecard and Source Excel files for demo use."""
    os.makedirs("sample_data", exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Create Scorecard template if not exists
    if not os.path.exists(SCORECARD_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = SCORECARD_SHEET_NAME
        ws["I4"] = "Metric Name"
        ws["J4"] = "Facility Values"
        ws["K4"] = "Region Values"
        for r in range(5, 22):
            ws.cell(row=r, column=9, value=f"Metric Row {r}")
        wb.save(SCORECARD_PATH)

    # Create Source data file if not exists (fake data)
    if not os.path.exists(SOURCE_PATH):
        wb = Workbook()
        wb.remove(wb.active)
        for sh in {"Sheet1", "Sheet2", "Sheet3"}:
            wb.create_sheet(sh)
        wb["Sheet1"]["B10"] = 1.23; wb["Sheet1"]["C10"] = 1.11
        wb["Sheet1"]["B16"] = 0.95; wb["Sheet1"]["C16"] = 0.90
        wb["Sheet2"]["B23"] = 0.88; wb["Sheet2"]["C23"] = 0.76
        wb["Sheet3"]["B25"] = 0.31; wb["Sheet3"]["C25"] = 0.29
        wb.save(SOURCE_PATH)

def main():
    _ensure_demo_files()

    wb_score = openpyxl.load_workbook(SCORECARD_PATH)
    ws_score = wb_score[SCORECARD_SHEET_NAME]
    wb_source = openpyxl.load_workbook(SOURCE_PATH, data_only=True)

    col_fac = column_index_from_string(FACILITY_COL_LETTER)
    col_reg = column_index_from_string(REGION_COL_LETTER)

    for metric in METRICS:
        sheet = metric["sheet"]
        fac_cell = metric["facility"]
        reg_cell = metric["region"]
        row_idx = metric["row"]
        mult = float(metric.get("multiply", 1))

        if sheet not in wb_source.sheetnames:
            raise KeyError(f"Missing sheet: {sheet}")

        ws_src = wb_source[sheet]
        val_fac = ws_src[fac_cell].value
        val_reg = ws_src[reg_cell].value

        val_fac = (val_fac or 0) * mult
        val_reg = (val_reg or 0) * mult

        ws_score.cell(row=row_idx, column=col_fac, value=val_fac)
        ws_score.cell(row=row_idx, column=col_reg, value=val_reg)

    out_path = os.path.join(OUTPUT_DIR, "Scorecard_filled.xlsx")
    wb_score.save(out_path)
    print(f"✅ Scorecard updated successfully: {out_path}")

if __name__ == "__main__":
    main()

